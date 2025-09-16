#!/usr/bin/env python3
"""
BOS Excel Dashboard Prototype Builder
Creates comprehensive workbook with:
- 5 normalized data sheets (CSV structure)
- 1 consolidated 52-field service view
- 1 dynamic dashboard
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
import numpy as np

def create_bos_workbook():
    """Create the complete BOS Excel workbook"""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create all worksheets - Entry forms first for better tab order
    po_entry_sheet = wb.create_sheet("PO_Entry_Form")
    dev_entry_sheet = wb.create_sheet("Dev_Entry_Form")
    ops_entry_sheet = wb.create_sheet("Ops_Entry_Form")
    dashboard_sheet = wb.create_sheet("Dashboard")
    service_model_sheet = wb.create_sheet("Service_Data_Model")
    services_sheet = wb.create_sheet("Services")
    sli_sheet = wb.create_sheet("SLI_Definitions")
    slo_sheet = wb.create_sheet("SLO_Configurations")
    impact_sheet = wb.create_sheet("Impact_Assessments")
    ops_sheet = wb.create_sheet("Operational_Metadata")
    
    # Create sample data matching CSV structure
    create_data_sheets(wb)
    create_entry_forms(wb)
    create_service_model_sheet(wb)
    create_dashboard_sheet(wb)
    
    return wb

def create_data_sheets(wb):
    """Create the 5 normalized data sheets with sample data"""
    
    # Services data
    services_data = [
        ["service_id", "serviceName", "displayName", "businessPurpose", "serviceType", "tierLevel", 
         "businessUnit", "performanceQuestion", "tags", "productOwner"],
        ["SVC001", "treasury-order-funding-service", "Treasury Order Funding Service", 
         "Execute wire transfers to complete home purchase transactions", "customer-facing", 1,
         "Home Lending", "What percentage of wire transfers complete successfully on scheduled closing date?",
         "mortgage", "sarah.chen@company.com"],
        ["SVC002", "credit-check-service", "Credit Check Service",
         "Validate borrower creditworthiness for loan approval decisions", "internal", 2,
         "Home Lending", "What percentage of credit checks return valid scores within acceptable time?",
         "lending;credit;validation", "michael.torres@company.com"]
    ]
    
    # SLI Definitions data
    sli_data = [
        ["service_id", "sliName", "sliDisplayName", "sliType", "goodEventsCriteria_PO", "goodEventsCriteria_Dev",
         "totalEventsCriteria_PO", "totalEventsCriteria_Dev", "thresholdQuery_Dev", "thresholdOperator", 
         "thresholdValue", "queryImplementation", "dataSource", "dataSourceDetails", "technicalOwner", "implementationNotes"],
        ["SVC001", "funding-success-rate", "Wire Transfer Success Rate", "ratioMetric",
         "Wire transfer completed successfully enabling scheduled closing",
         "status='FUNDED' AND closing_date=scheduled_date AND amount>0",
         "All wire transfer attempts for scheduled closings",
         "request_type='CLOSING_FUNDING' AND scheduled_date IS NOT NULL",
         "", "", "", "SELECT COUNT(*) WHERE status='FUNDED' AND closing_date=scheduled_date",
         "sql", '{"database":"loans","table":"wire_transfers"}', "platform-engineering-team",
         "Requires 5-minute aggregation window"],
        ["SVC002", "credit-check-success-rate", "Credit Check Success Rate", "ratioMetric",
         "Credit check returns valid score (300-850) within 5 seconds",
         "status='SUCCESS' AND score BETWEEN 300 AND 850 AND response_time_ms<5000",
         "All credit check requests for loan applications",
         "request_type='CREDIT_CHECK' AND source='LOAN_APP'",
         "", "", "", "SELECT COUNT(*) WHERE status='SUCCESS' AND score BETWEEN 300 AND 850",
         "sql", '{"database":"lending","table":"credit_checks"}', "platform-engineering-team",
         "Uses Equifax/Experian/TransUnion APIs with 5s timeout"]
    ]
    
    # SLO Configurations data
    slo_data = [
        ["service_id", "sloTarget", "sloTargetRationale", "timeWindow", "timeWindowType", "budgetingMethod",
         "timeSliceTarget", "timeSliceWindow", "alertingThreshold", "pageThreshold"],
        ["SVC001", 99.5, "Industry standard for critical financial transactions with manageable failure volume for escalation team",
         "7d", "rolling", "Occurrences", "", "", 99.3, 99.0],
        ["SVC002", 99.5, "Credit checks block loan approval; brief outages tolerable but impact customer experience",
         "7d", "rolling", "Occurrences", "", "", 99.3, 99.0]
    ]
    
    # Impact Assessments data (multiple per service)
    impact_data = [
        ["service_id", "impactCategory", "stakeholderType", "stakeholderCount", "failureScenario", "businessConsequence",
         "financialImpact", "regulatoryImpact", "customerImpactQuery", "financialImpactQuery", "legalRiskQuery", "operationalImpactQuery"],
        ["SVC001", "customer_experience", "homebuyers", "850 daily", "Wire transfer fails preventing scheduled closing",
         "Homebuyers cannot complete purchase and may lose rate lock", "", "",
         "COUNT(DISTINCT customer_id) WHERE status='FAILED'", "", "", ""],
        ["SVC001", "financial", "company", "", "Wire transfer failures prevent closings",
         "Lost transaction revenue and potential customer churn", "$380000 average per delayed closing", "",
         "", "SUM(wire_amount) WHERE status!='FUNDED'", "", ""],
        ["SVC002", "customer_experience", "loan applicants", "2000 daily", "Credit check fails preventing loan application progression",
         "Loan applicants stuck in process and may seek other lenders", "", "",
         "COUNT(DISTINCT applicant_id) WHERE status='FAILED'", "", "", ""],
        ["SVC002", "financial", "company", "", "Failed credit checks prevent loan origination",
         "Lost loan origination revenue opportunity", "$50000 average loan amount", "",
         "", "SUM(loan_amount) WHERE credit_check_status='FAILED'", "", ""],
        ["SVC002", "legal_risk", "company", "", "Credit checks exceed FCRA 30-day timing requirements",
         "Regulatory compliance violation risk", "", "Potential CFPB fines",
         "", "", "COUNT(*) WHERE response_time_days > 30", ""],
        ["SVC002", "operational", "loan processors", "50 processors", "Credit check failures require manual intervention",
         "Manual workarounds reduce processing capacity", "", "",
         "", "", "", "COUNT(*) WHERE requires_manual_review='TRUE'"]
    ]
    
    # Operational Metadata data
    ops_data = [
        ["service_id", "alertNotificationTargets", "dashboardUrl", "runbookUrl", "alertingConfigured", "lastValidated",
         "version", "created", "modified", "modifiedBy", "status", "reviewDate", "notes"],
        ["SVC001", "email:homelending-ops@company.com;pagerduty:PD-HL-001", "https://grafana.company.com/d/funding-slo",
         "https://wiki.company.com/runbooks/funding", True, "2024-01-15", "1.0", "2024-01-01T00:00:00Z",
         "2024-01-15T10:30:00Z", "sarah.chen@company.com", "active", "2024-07-01",
         "Initial BOS implementation for treasury service"],
        ["SVC002", "email:lending-platform@company.com;slack:#lending-alerts", "https://grafana.company.com/d/credit-check-slo",
         "https://wiki.company.com/credit-check-runbook", True, "2025-09-15", "1.0", "2025-09-15T18:00:00Z",
         "2025-09-15T18:00:00Z", "chad.johnson@company.com", "active", "2026-01-15",
         "Credit bureau integration service"]
    ]
    
    # Write data to sheets
    write_data_to_sheet(wb["Services"], services_data)
    write_data_to_sheet(wb["SLI_Definitions"], sli_data)
    write_data_to_sheet(wb["SLO_Configurations"], slo_data)
    write_data_to_sheet(wb["Impact_Assessments"], impact_data)
    write_data_to_sheet(wb["Operational_Metadata"], ops_data)

def write_data_to_sheet(sheet, data):
    """Write data array to worksheet with formatting"""
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Header formatting
            if row_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Border for all cells
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        sheet.column_dimensions[column_letter].width = adjusted_width

def create_entry_forms(wb):
    """Create the three persona-specific data entry forms"""
    
    # Product Owner Entry Form
    create_po_entry_form(wb)
    
    # Developer Entry Form
    create_dev_entry_form(wb)
    
    # Operations Entry Form
    create_ops_entry_form(wb)

def create_po_entry_form(wb):
    """Create Product Owner data entry form"""
    sheet = wb["PO_Entry_Form"]
    
    # Title
    sheet.merge_cells("A1:E1")
    title_cell = sheet["A1"]
    title_cell.value = "Product Owner Data Entry - Business Context & Requirements"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center")
    
    # Service selector
    sheet["A3"] = "Service to Edit:"
    sheet["B3"] = "SVC001"  # Default
    sheet["A3"].font = Font(bold=True)
    
    # Add data validation for service selection
    service_validation = DataValidation(type="list", formula1="Services!A2:A10")
    service_validation.add("B3")
    sheet.add_data_validation(service_validation)
    
    # Instructions
    sheet.merge_cells("A5:E5")
    sheet["A5"] = "Complete the fields below to define business context and success criteria for your service."
    sheet["A5"].font = Font(italic=True)
    
    current_row = 7
    
    # Service Definition Section
    po_service_fields = [
        ("Service ID", "Unique identifier for the service", "A", "readonly", '=B3'),
        ("Service Name", "Technical name used in systems (e.g., treasury-order-funding-service)", "B", "input", ""),
        ("Display Name", "Human-readable name for dashboards", "C", "input", ""),
        ("Business Purpose", "What the service does for the business in plain language", "D", "input", ""),
        ("Service Type", "One of: customer-facing, internal, infrastructure", "E", "dropdown", "customer-facing,internal,infrastructure"),
        ("Tier Level", "Criticality from 1 (most critical) to 6 (least critical)", "F", "input", ""),
        ("Business Unit", "Owning business organization", "G", "input", ""),
        ("Performance Question", "The key question this service's SLO answers", "H", "input", ""),
        ("Tags", "Comma-separated categorization labels", "I", "input", ""),
        ("Product Owner", "Email of the responsible Product Owner", "J", "input", "")
    ]
    
    current_row = add_form_section(sheet, "SERVICE DEFINITION (Services Table)", po_service_fields, current_row)
    current_row += 2
    
    # SLI Business Definition Section
    po_sli_fields = [
        ("SLI Name", "Technical identifier for the SLI", "B", "input", ""),
        ("SLI Display Name", "Human-readable name for dashboards", "C", "input", ""),
        ("Good Events (Business)", "Business language definition of success", "E", "input", ""),
        ("Total Events (Business)", "Business language definition of scope", "G", "input", "")
    ]
    
    current_row = add_form_section(sheet, "SLI BUSINESS DEFINITION (SLI_Definitions Table)", po_sli_fields, current_row)
    current_row += 2
    
    # SLO Configuration Section
    po_slo_fields = [
        ("SLO Target", "Target percentage (e.g., 99.5 for 99.5%)", "B", "input", ""),
        ("SLO Target Rationale", "Business justification for this target", "C", "input", ""),
        ("Time Window", "Measurement period (e.g., 7d, 28d, 1h)", "D", "input", "")
    ]
    
    current_row = add_form_section(sheet, "SLO TARGETS (SLO_Configurations Table)", po_slo_fields, current_row)
    current_row += 2
    
    # Impact Assessment Section
    impact_fields = [
        ("Impact Category", "Impact type: customer_experience, financial, legal_risk, operational", "B", "dropdown", "customer_experience,financial,legal_risk,operational"),
        ("Stakeholder Type", "Who is affected (for customer/operational impacts)", "C", "input", ""),
        ("Stakeholder Count", "Number affected (for customer/operational impacts)", "D", "input", ""),
        ("Failure Scenario", "Specific description of what failure looks like", "E", "input", ""),
        ("Business Consequence", "What happens when this fails", "F", "input", ""),
        ("Financial Impact", "Dollar amount (ONLY for financial impact rows)", "G", "input", ""),
        ("Regulatory Impact", "Compliance implications (ONLY for legal_risk rows)", "H", "input", "")
    ]
    
    add_form_section(sheet, "BUSINESS IMPACT ASSESSMENT (Impact_Assessments Table)", impact_fields, current_row)
    
    # Add save instructions
    sheet[f"A{current_row + 15}"] = "Instructions: Complete fields above, then manually copy values to corresponding data tables."
    sheet[f"A{current_row + 15}"].font = Font(italic=True, color="666666")
    
    # Column widths
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15

def create_dev_entry_form(wb):
    """Create Developer data entry form"""
    sheet = wb["Dev_Entry_Form"]
    
    # Title
    sheet.merge_cells("A1:E1")
    title_cell = sheet["A1"]
    title_cell.value = "Developer Data Entry - Technical Implementation"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center")
    
    # Service selector
    sheet["A3"] = "Service to Edit:"
    sheet["B3"] = "SVC001"  # Default
    sheet["A3"].font = Font(bold=True)
    
    # Add data validation
    service_validation = DataValidation(type="list", formula1="Services!A2:A10")
    service_validation.add("B3")
    sheet.add_data_validation(service_validation)
    
    # Instructions
    sheet.merge_cells("A5:E5")
    sheet["A5"] = "Define the technical implementation for measuring and monitoring this service."
    sheet["A5"].font = Font(italic=True)
    
    current_row = 7
    
    # Technical SLI Definition
    dev_sli_fields = [
        ("SLI Type", "One of: ratioMetric, thresholdMetric", "D", "dropdown", "ratioMetric,thresholdMetric"),
        ("Good Events (Technical)", "Technical query/condition for success", "F", "input", ""),
        ("Total Events (Technical)", "Technical query/condition for total events", "H", "input", ""),
        ("Threshold Query", "Query for threshold metrics (optional)", "I", "input", ""),
        ("Threshold Operator", "One of: lt, lte, gt, gte (optional)", "J", "dropdown", "lt,lte,gt,gte"),
        ("Threshold Value", "Numeric threshold value (optional)", "K", "input", ""),
        ("Query Implementation", "Actual production query details", "L", "input", ""),
        ("Data Source", "One of: sql, splunk, prometheus", "M", "dropdown", "sql,splunk,prometheus"),
        ("Data Source Details", "JSON with connection info", "N", "input", ""),
        ("Technical Owner", "Development team responsible", "O", "input", ""),
        ("Implementation Notes", "Technical context or special considerations", "P", "input", "")
    ]
    
    current_row = add_form_section(sheet, "TECHNICAL SLI IMPLEMENTATION (SLI_Definitions Table)", dev_sli_fields, current_row)
    current_row += 2
    
    # Impact Query Implementation
    impact_query_fields = [
        ("Customer Impact Query", "Query for customer impact measurement", "I", "input", ""),
        ("Financial Impact Query", "Query for financial impact measurement", "J", "input", ""),
        ("Legal Risk Query", "Query for legal/compliance risk measurement", "K", "input", ""),
        ("Operational Impact Query", "Query for operational impact measurement", "L", "input", "")
    ]
    
    add_form_section(sheet, "IMPACT MEASUREMENT QUERIES (Impact_Assessments Table)", impact_query_fields, current_row)
    
    # Column widths
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15

def create_ops_entry_form(wb):
    """Create Operations data entry form"""
    sheet = wb["Ops_Entry_Form"]
    
    # Title
    sheet.merge_cells("A1:E1")
    title_cell = sheet["A1"]
    title_cell.value = "Operations Data Entry - Deployment & Lifecycle"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center")
    
    # Service selector
    sheet["A3"] = "Service to Edit:"
    sheet["B3"] = "SVC001"  # Default
    sheet["A3"].font = Font(bold=True)
    
    # Add data validation
    service_validation = DataValidation(type="list", formula1="Services!A2:A10")
    service_validation.add("B3")
    sheet.add_data_validation(service_validation)
    
    # Instructions
    sheet.merge_cells("A5:E5")
    sheet["A5"] = "Configure operational metadata, alerting, and lifecycle management for this service."
    sheet["A5"].font = Font(italic=True)
    
    current_row = 7
    
    # Operational Metadata
    ops_fields = [
        ("Alert Notification Targets", "Semicolon-separated notification channels", "B", "input", ""),
        ("Dashboard URL", "Link to monitoring dashboard", "C", "input", ""),
        ("Runbook URL", "Link to incident response guide", "D", "input", ""),
        ("Alerting Configured", "Whether alerts are set up (true/false)", "E", "dropdown", "true,false"),
        ("Last Validated", "Date of last validation (YYYY-MM-DD)", "F", "input", ""),
        ("Version", "Record version number (e.g., 1.0, 1.1)", "G", "input", ""),
        ("Created", "Creation timestamp (ISO 8601 format)", "H", "input", ""),
        ("Modified", "Last modification timestamp (ISO 8601 format)", "I", "input", ""),
        ("Modified By", "Email of last person to modify", "J", "input", ""),
        ("Status", "Lifecycle state: draft, active, deprecated", "K", "dropdown", "draft,active,deprecated"),
        ("Review Date", "Next scheduled review date (YYYY-MM-DD)", "L", "input", ""),
        ("Notes", "General notes or comments", "M", "input", "")
    ]
    
    current_row = add_form_section(sheet, "OPERATIONAL METADATA (Operational_Metadata Table)", ops_fields, current_row)
    current_row += 2
    
    # SLO Operational Configuration
    slo_ops_fields = [
        ("Time Window Type", "One of: rolling, calendar", "E", "dropdown", "rolling,calendar"),
        ("Budgeting Method", "One of: Occurrences, Timeslices, RatioTimeslices", "F", "dropdown", "Occurrences,Timeslices,RatioTimeslices"),
        ("Time Slice Target", "Target for timeslice budgeting (optional, 0-1)", "G", "input", ""),
        ("Time Slice Window", "Duration of each time slice (optional, e.g., 1h)", "H", "input", ""),
        ("Alerting Threshold", "Warning threshold percentage", "I", "input", ""),
        ("Page Threshold", "Critical threshold percentage for paging", "J", "input", "")
    ]
    
    add_form_section(sheet, "SLO OPERATIONAL CONFIGURATION (SLO_Configurations Table)", slo_ops_fields, current_row)
    
    # Column widths
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15

def add_form_section(sheet, section_title, fields, start_row):
    """Add a form section with fields"""
    # Section header
    sheet.merge_cells(f"A{start_row}:E{start_row}")
    header_cell = sheet[f"A{start_row}"]
    header_cell.value = section_title
    header_cell.font = Font(size=12, bold=True, color="FFFFFF")
    header_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_cell.alignment = Alignment(horizontal="center")
    
    current_row = start_row + 2
    
    # Column headers
    sheet[f"A{current_row}"] = "Field Name"
    sheet[f"B{current_row}"] = "Value"
    sheet[f"C{current_row}"] = "Description"
    sheet[f"D{current_row}"] = "Table Column"
    sheet[f"E{current_row}"] = "Type"
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        cell = sheet[f"{col}{current_row}"]
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    current_row += 1
    
    # Add fields
    for field_name, description, column_ref, field_type, default_value in fields:
        sheet[f"A{current_row}"] = field_name
        sheet[f"C{current_row}"] = description
        sheet[f"D{current_row}"] = column_ref
        sheet[f"E{current_row}"] = field_type
        
        # Value field
        value_cell = sheet[f"B{current_row}"]
        if field_type == "dropdown" and default_value:
            # Add data validation for dropdowns
            validation = DataValidation(type="list", formula1=f'"{default_value}"')
            validation.add(f"B{current_row}")
            sheet.add_data_validation(validation)
        elif field_type == "readonly":
            value_cell.value = default_value
            value_cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Styling
        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = sheet[f"{col}{current_row}"]
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            if col == 'A':
                cell.font = Font(bold=True)
        
        current_row += 1
    
    return current_row

def create_service_model_sheet(wb):
    sheet = wb["Service_Data_Model"]
    
    # Title
    sheet.merge_cells("A1:D1")
    title_cell = sheet["A1"]
    title_cell.value = "BOS Service Data Model - Complete 52-Field Profile"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center")
    
    # Service selector with display names
    sheet["A3"] = "Selected Service:"
    sheet["B3"] = "Treasury Order Funding Service"  # Default display name
    
    # Add data validation for service selection using display names
    service_validation = DataValidation(type="list", formula1="Services!C2:C10")
    service_validation.add("B3")
    sheet.add_data_validation(service_validation)
    
    # Helper cell to convert display name back to service_id for lookups
    sheet["D3"] = '=INDEX(Services!A:A,MATCH(B3,Services!C:C,0))'  # Hidden lookup helper
    
    # Section headers and field layout
    current_row = 5
    
    # Product Owner Section (26 fields)
    sheet.merge_cells(f"A{current_row}:D{current_row}")
    po_header = sheet[f"A{current_row}"]
    po_header.value = "PRODUCT OWNER FIELDS (22 fields - Business Context & Requirements)"
    po_header.font = Font(size=12, bold=True, color="FFFFFF")
    po_header.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    po_header.alignment = Alignment(horizontal="center")
    current_row += 2
    
    # Column headers for field display
    sheet[f"A{current_row}"] = "Field Name"
    sheet[f"B{current_row}"] = "Value"
    sheet[f"C{current_row}"] = "Description"
    
    for col in ['A', 'B', 'C']:
        cell = sheet[f"{col}{current_row}"]
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    current_row += 1
    
    # PO fields with formulas and descriptions
    po_fields = [
        ("Service ID", '=IF(D3="","",INDEX(Services!A:A,MATCH(D3,Services!A:A,0)))', "Unique identifier for the service"),
        ("Service Name", '=IF(D3="","",INDEX(Services!B:B,MATCH(D3,Services!A:A,0)))', "Technical name used in systems"),
        ("Display Name", '=IF(D3="","",INDEX(Services!C:C,MATCH(D3,Services!A:A,0)))', "Human-readable name for dashboards"),
        ("Business Purpose", '=IF(D3="","",INDEX(Services!D:D,MATCH(D3,Services!A:A,0)))', "What the service does for the business in plain language"),
        ("Service Type", '=IF(D3="","",INDEX(Services!E:E,MATCH(D3,Services!A:A,0)))', "One of: customer-facing, internal, infrastructure"),
        ("Tier Level", '=IF(D3="","",INDEX(Services!F:F,MATCH(D3,Services!A:A,0)))', "Criticality from 1 (most critical) to 6 (least critical)"),
        ("Business Unit", '=IF(D3="","",INDEX(Services!G:G,MATCH(D3,Services!A:A,0)))', "Owning business organization"),
        ("Performance Question", '=IF(D3="","",INDEX(Services!H:H,MATCH(D3,Services!A:A,0)))', "The key question this service's SLO answers"),
        ("Tags", '=IF(D3="","",INDEX(Services!I:I,MATCH(D3,Services!A:A,0)))', "Comma-separated categorization labels"),
        ("Product Owner", '=IF(D3="","",INDEX(Services!J:J,MATCH(D3,Services!A:A,0)))', "Email of the responsible Product Owner"),
        ("SLI Name", '=IF(D3="","",INDEX(SLI_Definitions!B:B,MATCH(D3,SLI_Definitions!A:A,0)))', "Technical identifier for the SLI"),
        ("SLI Display Name", '=IF(D3="","",INDEX(SLI_Definitions!C:C,MATCH(D3,SLI_Definitions!A:A,0)))', "Human-readable name for dashboards"),
        ("Good Events (Business)", '=IF(D3="","",INDEX(SLI_Definitions!E:E,MATCH(D3,SLI_Definitions!A:A,0)))', "Business language definition of success"),
        ("Total Events (Business)", '=IF(D3="","",INDEX(SLI_Definitions!G:G,MATCH(D3,SLI_Definitions!A:A,0)))', "Business language definition of scope"),
        ("SLO Target", '=IF(D3="","",INDEX(SLO_Configurations!B:B,MATCH(D3,SLO_Configurations!A:A,0)))', "Target percentage (e.g., 99.5 for 99.5%)"),
        ("SLO Rationale", '=IF(D3="","",INDEX(SLO_Configurations!C:C,MATCH(D3,SLO_Configurations!A:A,0)))', "Business justification for this target"),
        ("Time Window", '=IF(D3="","",INDEX(SLO_Configurations!D:D,MATCH(D3,SLO_Configurations!A:A,0)))', "Measurement period (e.g., 7d, 28d, 1h)"),
        ("Impact Category", '=IF(D3="","",INDEX(Impact_Assessments!B:B,MATCH(D3,Impact_Assessments!A:A,0)))', "Impact type: customer_experience, financial, legal_risk, operational"),
        ("Stakeholder Type", '=IF(D3="","",INDEX(Impact_Assessments!C:C,MATCH(D3,Impact_Assessments!A:A,0)))', "Who is affected"),
        ("Stakeholder Count", '=IF(D3="","",INDEX(Impact_Assessments!D:D,MATCH(D3,Impact_Assessments!A:A,0)))', "Number affected"),
        ("Failure Scenario", '=IF(D3="","",INDEX(Impact_Assessments!E:E,MATCH(D3,Impact_Assessments!A:A,0)))', "Specific description of what failure looks like"),
        ("Business Consequence", '=IF(D3="","",INDEX(Impact_Assessments!F:F,MATCH(D3,Impact_Assessments!A:A,0)))', "What happens when this fails"),
        ("Financial Impact", '=IF(D3="","",INDEX(Impact_Assessments!G:G,MATCH(D3,Impact_Assessments!A:A,0)))', "Dollar amount (for financial impact rows)"),
        ("Regulatory Impact", '=IF(D3="","",INDEX(Impact_Assessments!H:H,MATCH(D3,Impact_Assessments!A:A,0)))', "Compliance implications (for legal_risk rows)")
    ]
    
    current_row = add_model_field_section(sheet, po_fields, current_row, "E8F5E8")  # Light green
    current_row += 2
    
    # Developer Section (16 fields)
    sheet.merge_cells(f"A{current_row}:D{current_row}")
    dev_header = sheet[f"A{current_row}"]
    dev_header.value = "DEVELOPER FIELDS (15 fields - Technical Implementation)"
    dev_header.font = Font(size=12, bold=True, color="FFFFFF")
    dev_header.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    dev_header.alignment = Alignment(horizontal="center")
    current_row += 2
    
    # Developer fields
    dev_fields = [
        ("SLI Type", '=IF(D3="","",INDEX(SLI_Definitions!D:D,MATCH(D3,SLI_Definitions!A:A,0)))', "One of: ratioMetric, thresholdMetric"),
        ("Good Events (Technical)", '=IF(D3="","",INDEX(SLI_Definitions!F:F,MATCH(D3,SLI_Definitions!A:A,0)))', "Technical query/condition for success"),
        ("Total Events (Technical)", '=IF(D3="","",INDEX(SLI_Definitions!H:H,MATCH(D3,SLI_Definitions!A:A,0)))', "Technical query/condition for total events"),
        ("Threshold Query", '=IF(D3="","",INDEX(SLI_Definitions!I:I,MATCH(D3,SLI_Definitions!A:A,0)))', "Query for threshold metrics (optional)"),
        ("Threshold Operator", '=IF(D3="","",INDEX(SLI_Definitions!J:J,MATCH(D3,SLI_Definitions!A:A,0)))', "One of: lt, lte, gt, gte (optional)"),
        ("Threshold Value", '=IF(D3="","",INDEX(SLI_Definitions!K:K,MATCH(D3,SLI_Definitions!A:A,0)))', "Numeric threshold value (optional)"),
        ("Query Implementation", '=IF(D3="","",INDEX(SLI_Definitions!L:L,MATCH(D3,SLI_Definitions!A:A,0)))', "Actual production query details"),
        ("Data Source", '=IF(D3="","",INDEX(SLI_Definitions!M:M,MATCH(D3,SLI_Definitions!A:A,0)))', "One of: sql, splunk, prometheus"),
        ("Data Source Details", '=IF(D3="","",INDEX(SLI_Definitions!N:N,MATCH(D3,SLI_Definitions!A:A,0)))', "JSON with connection info"),
        ("Technical Owner", '=IF(D3="","",INDEX(SLI_Definitions!O:O,MATCH(D3,SLI_Definitions!A:A,0)))', "Development team responsible"),
        ("Implementation Notes", '=IF(D3="","",INDEX(SLI_Definitions!P:P,MATCH(D3,SLI_Definitions!A:A,0)))', "Technical context or special considerations"),
        ("Customer Impact Query", '=IF(D3="","",INDEX(Impact_Assessments!I:I,MATCH(D3,Impact_Assessments!A:A,0)))', "Query for customer impact measurement"),
        ("Financial Impact Query", '=IF(D3="","",INDEX(Impact_Assessments!J:J,MATCH(D3,Impact_Assessments!A:A,0)))', "Query for financial impact measurement"),
        ("Legal Risk Query", '=IF(D3="","",INDEX(Impact_Assessments!K:K,MATCH(D3,Impact_Assessments!A:A,0)))', "Query for legal/compliance risk measurement"),
        ("Operational Impact Query", '=IF(D3="","",INDEX(Impact_Assessments!L:L,MATCH(D3,Impact_Assessments!A:A,0)))', "Query for operational impact measurement")
    ]
    
    current_row = add_model_field_section(sheet, dev_fields, current_row, "E1F4FD")  # Light blue
    current_row += 2
    
    # Operations Section (10 fields)
    sheet.merge_cells(f"A{current_row}:D{current_row}")
    ops_header = sheet[f"A{current_row}"]
    ops_header.value = "OPERATIONS FIELDS (16 fields - Deployment & Lifecycle)"
    ops_header.font = Font(size=12, bold=True, color="FFFFFF")
    ops_header.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ops_header.alignment = Alignment(horizontal="center")
    current_row += 2
    
    # Operations fields
    ops_fields = [
        ("Alert Notification Targets", '=IF(D3="","",INDEX(Operational_Metadata!B:B,MATCH(D3,Operational_Metadata!A:A,0)))', "Semicolon-separated notification channels"),
        ("Dashboard URL", '=IF(D3="","",INDEX(Operational_Metadata!C:C,MATCH(D3,Operational_Metadata!A:A,0)))', "Link to monitoring dashboard"),
        ("Runbook URL", '=IF(D3="","",INDEX(Operational_Metadata!D:D,MATCH(D3,Operational_Metadata!A:A,0)))', "Link to incident response guide"),
        ("Alerting Configured", '=IF(D3="","",INDEX(Operational_Metadata!E:E,MATCH(D3,Operational_Metadata!A:A,0)))', "Whether alerts are set up (true/false)"),
        ("Last Validated", '=IF(D3="","",INDEX(Operational_Metadata!F:F,MATCH(D3,Operational_Metadata!A:A,0)))', "Date of last validation (YYYY-MM-DD)"),
        ("Version", '=IF(D3="","",INDEX(Operational_Metadata!G:G,MATCH(D3,Operational_Metadata!A:A,0)))', "Record version number (e.g., 1.0, 1.1)"),
        ("Created", '=IF(D3="","",INDEX(Operational_Metadata!H:H,MATCH(D3,Operational_Metadata!A:A,0)))', "Creation timestamp (ISO 8601 format)"),
        ("Modified", '=IF(D3="","",INDEX(Operational_Metadata!I:I,MATCH(D3,Operational_Metadata!A:A,0)))', "Last modification timestamp (ISO 8601 format)"),
        ("Modified By", '=IF(D3="","",INDEX(Operational_Metadata!J:J,MATCH(D3,Operational_Metadata!A:A,0)))', "Email of last person to modify"),
        ("Status", '=IF(D3="","",INDEX(Operational_Metadata!K:K,MATCH(D3,Operational_Metadata!A:A,0)))', "Lifecycle state: draft, active, deprecated"),
        ("Review Date", '=IF(D3="","",INDEX(Operational_Metadata!L:L,MATCH(D3,Operational_Metadata!A:A,0)))', "Next scheduled review date (YYYY-MM-DD)"),
        ("Notes", '=IF(D3="","",INDEX(Operational_Metadata!M:M,MATCH(D3,Operational_Metadata!A:A,0)))', "General notes or comments"),
        ("Time Window Type", '=IF(D3="","",INDEX(SLO_Configurations!E:E,MATCH(D3,SLO_Configurations!A:A,0)))', "One of: rolling, calendar"),
        ("Budgeting Method", '=IF(D3="","",INDEX(SLO_Configurations!F:F,MATCH(D3,SLO_Configurations!A:A,0)))', "One of: Occurrences, Timeslices, RatioTimeslices"),
        ("Time Slice Target", '=IF(D3="","",INDEX(SLO_Configurations!G:G,MATCH(D3,SLO_Configurations!A:A,0)))', "Target for timeslice budgeting (optional, 0-1)"),
        ("Time Slice Window", '=IF(D3="","",INDEX(SLO_Configurations!H:H,MATCH(D3,SLO_Configurations!A:A,0)))', "Duration of each time slice (optional)"),
        ("Alerting Threshold", '=IF(D3="","",INDEX(SLO_Configurations!I:I,MATCH(D3,SLO_Configurations!A:A,0)))', "Warning threshold percentage"),
        ("Page Threshold", '=IF(D3="","",INDEX(SLO_Configurations!J:J,MATCH(D3,SLO_Configurations!A:A,0)))', "Critical threshold percentage for paging")
    ]
    
    add_model_field_section(sheet, ops_fields, current_row, "F2F2F2")  # Light gray

def add_model_field_section(sheet, fields, start_row, bg_color):
    """Add a section of fields to the service model sheet with color coding"""
    current_row = start_row
    for field_name, formula, description in fields:
        sheet[f"A{current_row}"] = field_name
        sheet[f"A{current_row}"].font = Font(bold=True)
        sheet[f"B{current_row}"] = formula
        sheet[f"C{current_row}"] = description
        
        # Color coding by persona
        for col in ['A', 'B', 'C']:
            cell = sheet[f"{col}{current_row}"]
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        current_row += 1
    
    # Adjust column widths
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 50
    
    return current_row

def add_field_section(sheet, fields, start_row):
    """Add a section of fields to the sheet"""
    current_row = start_row
    for field_name, formula in fields:
        sheet[f"A{current_row}"] = field_name
        sheet[f"A{current_row}"].font = Font(bold=True)
        sheet[f"B{current_row}"] = formula
        
        # Styling
        sheet[f"A{current_row}"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        sheet[f"A{current_row}"].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        sheet[f"B{current_row}"].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        current_row += 1
    
    # Adjust column widths
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 60
    
    return current_row

def create_dashboard_sheet(wb):
    """Create an enhanced professional dashboard"""
    sheet = wb["Dashboard"]
    
    # Title with professional styling
    sheet.merge_cells("A1:I1")
    title_cell = sheet["A1"]
    title_cell.value = "Business Observability Service Dashboard"
    title_cell.font = Font(size=18, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="2F5597", end_color="4472C4", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set row height for title
    sheet.row_dimensions[1].height = 30
    
    # Service selector with enhanced styling
    sheet["A3"] = "Service:"
    sheet["A3"].font = Font(size=12, bold=True)
    sheet["B3"] = "Treasury Order Funding Service"  # Default display name
    sheet["B3"].font = Font(size=12)
    
    # Style the service selector
    sheet["A3"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sheet["B3"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    sheet["B3"].border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Add data validation for service selection using display names
    service_validation = DataValidation(type="list", formula1="Services!C2:C10")
    service_validation.add("B3")
    sheet.add_data_validation(service_validation)
    
    # Helper cell to convert display name back to service_id for lookups
    sheet["A1"] = '=INDEX(Services!A:A,MATCH(B3,Services!C:C,0))'  # Hidden lookup helper
    
    # SERVICE CONTEXT SECTION
    create_professional_section_header(sheet, "SERVICE CONTEXT", 5, "70AD47")
    
    # Service info in professional two-column layout
    sheet["A7"] = "Service Name:"
    sheet["B7"] = '=IFERROR(INDEX(Services!B:B,MATCH(A1,Services!A:A,0)), "Not Defined")'
    sheet["E7"] = "Tier Level:"
    sheet["F7"] = '=IFERROR(INDEX(Services!F:F,MATCH(A1,Services!A:A,0)), "Not Defined")'
    
    sheet["A8"] = "Business Purpose:"
    sheet.merge_cells("B8:I8")
    sheet["B8"] = '=IFERROR(INDEX(Services!D:D,MATCH(A1,Services!A:A,0)), "Not Defined")'
    sheet["B8"].alignment = Alignment(wrap_text=True, vertical="top")
    
    sheet["A9"] = "Performance Question:"
    sheet.merge_cells("B9:I9")
    sheet["B9"] = '=IFERROR(INDEX(Services!H:H,MATCH(A1,Services!A:A,0)), "Not Defined")'
    sheet["B9"].alignment = Alignment(wrap_text=True, vertical="top")
    
    # Style service context section
    for row in [7, 8, 9]:
        sheet[f"A{row}"].font = Font(bold=True, size=10)
        sheet[f"A{row}"].fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        if row == 8 or row == 9:
            sheet[f"B{row}"].fill = PatternFill(start_color="F8FFF8", end_color="F8FFF8", fill_type="solid")
    
    sheet[f"E7"].font = Font(bold=True, size=10)
    sheet[f"E7"].fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
    
    # SERVICE LEVEL INDICATORS SECTION
    create_professional_section_header(sheet, "SERVICE LEVEL INDICATORS", 12, "5B9BD5")
    
    # SLI Name with enhanced styling
    sheet["A13"] = "SLI NAME:"
    sheet.merge_cells("B13:I13")
    sheet["B13"] = '=IFERROR(INDEX(SLI_Definitions!C:C,MATCH(A1,SLI_Definitions!A:A,0)), "Not Defined")'
    sheet["A13"].font = Font(size=11, bold=True)
    sheet["B13"].font = Font(size=12, bold=True, color="2F5597")
    sheet["A13"].fill = PatternFill(start_color="E1F4FD", end_color="E1F4FD", fill_type="solid")
    sheet["B13"].fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    
    # Enhanced stats boxes
    create_enhanced_stats_boxes(sheet, 15)
    
    # SLI Details with professional formatting
    sli_details = [
        ("Good Events:", '=IFERROR(INDEX(SLI_Definitions!E:E,MATCH(A1,SLI_Definitions!A:A,0)), "Not Defined")'),
        ("Total Events:", '=IFERROR(INDEX(SLI_Definitions!G:G,MATCH(A1,SLI_Definitions!A:A,0)), "Not Defined")'),
        ("Technical Query:", '=IFERROR(INDEX(SLI_Definitions!F:F,MATCH(A1,SLI_Definitions!A:A,0)), "Not Defined")')
    ]
    
    for i, (label, formula) in enumerate(sli_details):
        row = 19 + i
        sheet[f"A{row}"] = label
        sheet.merge_cells(f"B{row}:I{row}")
        sheet[f"B{row}"] = formula
        
        # Style SLI details
        sheet[f"A{row}"].font = Font(bold=True, size=10)
        sheet[f"A{row}"].fill = PatternFill(start_color="E1F4FD", end_color="E1F4FD", fill_type="solid")
        sheet[f"B{row}"].fill = PatternFill(start_color="F8FCFF", end_color="F8FCFF", fill_type="solid")
        sheet[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    
    # BUSINESS IMPACT & OWNERSHIP SECTION
    create_professional_section_header(sheet, "BUSINESS IMPACT & OWNERSHIP", 24, "C5504B")
    
    # Impact details with enhanced styling
    sheet["A25"] = "WHEN THIS FAILS:"
    sheet["A25"].font = Font(bold=True, size=11, color="C5504B")
    sheet["A25"].fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    
    impact_details = [
        ("Scenario:", '=IFERROR(INDEX(Impact_Assessments!E:E,MATCH(A1,Impact_Assessments!A:A,0)), "Not Defined")'),
        ("Impact:", '=IFERROR(INDEX(Impact_Assessments!F:F,MATCH(A1,Impact_Assessments!A:A,0)), "Not Defined")')
    ]
    
    for i, (label, formula) in enumerate(impact_details):
        row = 26 + i
        sheet[f"A{row}"] = label
        sheet.merge_cells(f"B{row}:I{row}")
        sheet[f"B{row}"] = formula
        
        # Style impact details
        sheet[f"A{row}"].font = Font(bold=True, size=10)
        sheet[f"A{row}"].fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
        sheet[f"B{row}"].fill = PatternFill(start_color="FFF5F8", end_color="FFF5F8", fill_type="solid")
        sheet[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    
    # Affected stakeholders with professional layout
    sheet["A28"] = "Affected:"
    sheet["B28"] = '=IFERROR(INDEX(Impact_Assessments!D:D,MATCH(A1,Impact_Assessments!A:A,0)), "Not Defined")'
    sheet["C28"] = '=IFERROR(INDEX(Impact_Assessments!C:C,MATCH(A1,Impact_Assessments!A:A,0)), "Not Defined")'
    
    sheet["E28"] = "Financial:"
    sheet.merge_cells("F28:I28")
    sheet["F28"] = '=IFERROR(INDEX(Impact_Assessments!G:G,MATCH(A1,Impact_Assessments!A:A,0)), "Not Defined")'
    
    # Style stakeholder info
    for col in ['A', 'E']:
        sheet[f"{col}28"].font = Font(bold=True, size=10)
        sheet[f"{col}28"].fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    
    # OWNERSHIP & SERVICE CONTEXT (Two columns)
    ownership_row = 30
    
    # Left column - OWNERSHIP
    sheet[f"A{ownership_row}"] = "OWNERSHIP:"
    sheet[f"A{ownership_row}"].font = Font(bold=True, size=11, color="2F5597")
    sheet[f"A{ownership_row}"].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    # Right column - SERVICE CONTEXT  
    sheet[f"E{ownership_row}"] = "SERVICE CONTEXT:"
    sheet[f"E{ownership_row}"].font = Font(bold=True, size=11, color="2F5597")
    sheet[f"E{ownership_row}"].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    ownership_fields = [
        ("Product Owner:", '=IFERROR(INDEX(Services!J:J,MATCH(A1,Services!A:A,0)), "Not Defined")', "Type:", '=IFERROR(INDEX(Services!E:E,MATCH(A1,Services!A:A,0)), "Not Defined")'),
        ("Technical Owner:", '=IFERROR(INDEX(SLI_Definitions!O:O,MATCH(A1,SLI_Definitions!A:A,0)), "Not Defined")', "Business Unit:", '=IFERROR(INDEX(Services!G:G,MATCH(A1,Services!A:A,0)), "Not Defined")'),
        ("Status:", '=IFERROR(INDEX(Operational_Metadata!K:K,MATCH(A1,Operational_Metadata!A:A,0)), "Not Defined")', "Service ID:", '=IFERROR(INDEX(Services!A:A,MATCH(A1,Services!A:A,0)), "Not Defined")')
    ]
    
    for i, (left_label, left_formula, right_label, right_formula) in enumerate(ownership_fields):
        row = ownership_row + 1 + i
        
        # Left side
        sheet[f"A{row}"] = left_label
        sheet[f"B{row}"] = left_formula
        sheet[f"A{row}"].font = Font(bold=True, size=10)
        sheet[f"A{row}"].fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        
        # Right side
        sheet[f"E{row}"] = right_label
        sheet[f"F{row}"] = right_formula
        sheet[f"E{row}"].font = Font(bold=True, size=10)
        sheet[f"E{row}"].fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    
    # Set professional column widths
    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 18
    sheet.column_dimensions['F'].width = 25
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 15
    sheet.column_dimensions['I'].width = 15

def create_professional_section_header(sheet, title, row, color):
    """Create a professional section header"""
    sheet.merge_cells(f"A{row}:I{row}")
    header_cell = sheet[f"A{row}"]
    header_cell.value = title
    header_cell.font = Font(size=12, bold=True, color="FFFFFF")
    header_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[row].height = 25

def create_enhanced_stats_boxes(sheet, row):
    """Create professional, dynamic stats indicator boxes"""
    stats_data = [
        ("CURRENT", '=IF(A1="SVC001","99.2%",IF(A1="SVC002","97.8%","No Data"))', "2F5597"),
        ("TARGET", '=IFERROR(INDEX(SLO_Configurations!B:B,MATCH(A1,SLO_Configurations!A:A,0)) & "%", "No Target")', "2F5597"),
        ("STATUS", '=IF(A1="SVC001","⚠️ WARNING",IF(A1="SVC002","✅ OK","Unknown"))', "2F5597"),
        ("TREND", '"📊 Stable"', "2F5597")
    ]
    
    # Create stats boxes in columns A, C, E, G
    cols = ['A', 'C', 'E', 'G']
    
    for i, (header, formula, color) in enumerate(stats_data):
        col = cols[i]
        
        # Header cell
        header_cell = sheet[f"{col}{row}"]
        header_cell.value = header
        header_cell.font = Font(size=10, bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cell.border = Border(
            left=Side(style='thick'), right=Side(style='thick'),
            top=Side(style='thick'), bottom=Side(style='thin')
        )
        
        # Value cell
        value_cell = sheet[f"{col}{row+1}"]
        value_cell.value = formula
        value_cell.font = Font(size=14, bold=True)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.border = Border(
            left=Side(style='thick'), right=Side(style='thick'),
            top=Side(style='thin'), bottom=Side(style='thick')
        )
        
        # Special formatting for status column
        if header == "STATUS":
            value_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Set row heights for stats boxes
    sheet.row_dimensions[row].height = 20
    sheet.row_dimensions[row+1].height = 35



def create_stats_boxes(sheet, row):
    """Create the stats indicator boxes"""
    # Headers
    sheet[f"A{row}"] = "CURRENT"
    sheet[f"C{row}"] = "TARGET"
    sheet[f"E{row}"] = "STATUS"
    sheet[f"G{row}"] = "TREND"
    
    # Values with conditional logic for different services using helper cell A1
    sheet[f"A{row+1}"] = '=IF(A1="SVC001","99.2%",IF(A1="SVC002","97.8%",""))'
    sheet[f"C{row+1}"] = '=IF(A1="","",INDEX(SLO_Configurations!B:B,MATCH(A1,SLO_Configurations!A:A,0)) & "%")'
    sheet[f"E{row+1}"] = '=IF(A1="SVC001","⚠️ WARNING",IF(A1="SVC002","✅ OK",""))'
    sheet[f"G{row+1}"] = "▄▃▅▆▇▆▅▄▃▅▆▇▆▅"  # Simple ASCII trend
    
    # Styling for stats boxes
    for col in ['A', 'C', 'E', 'G']:
        header_cell = sheet[f"{col}{row}"]
        value_cell = sheet[f"{col}{row+1}"]
        
        # Header styling
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center")
        header_cell.border = Border(
            left=Side(style='thick'), right=Side(style='thick'),
            top=Side(style='thick'), bottom=Side(style='thin')
        )
        
        # Value styling
        value_cell.font = Font(size=14, bold=True)
        value_cell.alignment = Alignment(horizontal="center")
        value_cell.border = Border(
            left=Side(style='thick'), right=Side(style='thick'),
            top=Side(style='thin'), bottom=Side(style='thick')
        )
        
        # Conditional coloring for status with accessibility
        if col == 'E':  # Status column
            value_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

def apply_dashboard_formatting(sheet):
    """Apply consistent formatting to dashboard"""
    # Set default font
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and not cell.font.bold:
                cell.font = Font(name='Calibri', size=10)
    
    # Field labels
    label_cells = ['A7', 'E7', 'A8', 'A9', 'A18', 'A19', 'A20', 'A25', 'A26', 'A27', 'E27', 'A30', 'E30', 'A31', 'E31', 'A32', 'E32']
    for cell_ref in label_cells:
        cell = sheet[cell_ref]
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Main execution
if __name__ == "__main__":
    print("Building BOS Excel Dashboard Prototype...")
    workbook = create_bos_workbook()
    
    # Save to outputs directory
    output_path = "/mnt/user-data/outputs/BOS_Dashboard_Prototype_v3.4.xlsx"
    workbook.save(output_path)
    print(f"Workbook saved to: {output_path}")
    print("\nWorkbook contains:")
    print("- PO_Entry_Form: Product Owner data entry (22 fields)")
    print("- Dev_Entry_Form: Developer data entry (15 fields)")
    print("- Ops_Entry_Form: Operations data entry (18 fields)")
    print("- Dashboard: Dynamic service dashboard with stats and visualizations")
    print("- Service_Data_Model: Complete 53-field profile with descriptions and color coding")
    print("- Services: Core service definitions")
    print("- SLI_Definitions: Technical measurement specifications")
    print("- SLO_Configurations: Performance targets and thresholds")
    print("- Impact_Assessments: Business impact scenarios")
    print("- Operational_Metadata: Deployment and lifecycle information")
    print("\nKey Features:")
    print("✓ Service dropdowns use display names")
    print("✓ Persona fields color-coded (green=PO, blue=Dev, gray=Ops)")
    print("✓ Field descriptions visible alongside values")
    print("✓ Accessible status indicators (icons + text)")
    print("✓ Complete data entry workflows for all personas")
    print("✓ CORRECTED field ownership based on CSV documentation")
    print("✓ FIXED Service_Data_Model formula references (PO fields now populate correctly)")
    print("✓ ENHANCED dashboard with professional styling and improved UX")
    print("✓ Dynamic error handling with IFERROR functions")
    print("✓ Responsive layout with proper text wrapping and column widths")
    print("✓ Professional section headers and color-coded backgrounds")
    print("✓ Enhanced stats boxes with better visual hierarchy")
    print("✓ FIXED file corruption issue (removed problematic conditional formatting)")
